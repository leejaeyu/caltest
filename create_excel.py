import openpyxl

# 새 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active

# 데이터 입력
ws['A1'] = '이름'
ws['B1'] = '나이'
ws['A2'] = '철수'
ws['B2'] = 30
ws['A3'] = '영희'
ws['B3'] = 28

# 파일 저장
wb.save('sample_data.xlsx')
